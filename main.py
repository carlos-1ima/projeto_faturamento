import pandas as pd
import os
from datetime import datetime

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# CONFIGURAÇÃO
pasta = "entrada"
saida = "saida"

if not os.path.exists(saida):
    os.makedirs(saida)

df_total = pd.DataFrame()

# FUNÇÃO DE LEITURA
def processar_linha(linha):
    partes = linha.split(";")
    
    try:
        competencia = partes[2]
        situacao = partes[4].strip().upper()
        cnpj = partes[6]
        empresa = partes[8]
        valor = partes[12]
        
        if situacao != "NORMAL":
            return None
        
        valor = valor.replace(".", "").replace(",", ".")
        valor = float(valor)
        
        return {
            "CNPJ": cnpj,
            "Empresa": empresa,
            "Competencia": competencia,
            "Valor": valor
        }
    except:
        return None

# LEITURA DOS CSVs
for arquivo in os.listdir(pasta):
    if arquivo.endswith(".csv"):
        caminho = os.path.join(pasta, arquivo)
        
        print(f"Processando: {arquivo}")
        
        with open(caminho, encoding="latin1") as f:
            linhas = f.readlines()
        
        for linha in linhas[1:]:
            resultado = processar_linha(linha)
            if resultado:
                df_total = pd.concat([df_total, pd.DataFrame([resultado])], ignore_index=True)

# AGRUPAMENTO
faturamento = (
    df_total
    .groupby(["CNPJ", "Empresa", "Competencia"])["Valor"]
    .sum()
    .reset_index()
)

# PIS / COFINS
faturamento["PIS"] = faturamento["Valor"] * 0.0065
faturamento["COFINS"] = faturamento["Valor"] * 0.03

# TRIMESTRE
def get_trimestre(mes):
    mes_num = int(mes.split("/")[0])
    
    if mes_num <= 3:
        return "1T"
    elif mes_num <= 6:
        return "2T"
    elif mes_num <= 9:
        return "3T"
    else:
        return "4T"

faturamento["Trimestre"] = faturamento["Competencia"].apply(get_trimestre)

# TOTAL TRIMESTRE
faturamento["Total_Trimestre"] = (
    faturamento
    .groupby(["CNPJ", "Trimestre"])["Valor"]
    .transform("sum")
    .round(2)
)

# IRPJ / CSLL
faturamento["IRPJ"] = 0.0
faturamento["CSLL"] = 0.0

for (cnpj, trimestre), grupo in faturamento.groupby(["CNPJ", "Trimestre"]):
    
    faturamento_trimestre = grupo["Valor"].sum()
    
    base_irpj = faturamento_trimestre * 0.16
    base_csll = faturamento_trimestre * 0.32
    
    irpj = base_irpj * 0.15
    csll = base_csll * 0.09
    
    if base_irpj > 60000:
        adicional = (base_irpj - 60000) * 0.10
        irpj += adicional
    
    idx = grupo.index[-1]
    
    faturamento.loc[idx, "IRPJ"] = irpj
    faturamento.loc[idx, "CSLL"] = csll

# EXPORTAÇÃO
nome_arquivo = f"faturamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
arquivo_saida = os.path.join(saida, nome_arquivo)

with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    
    for mes in faturamento["Competencia"].unique():
        df_mes = faturamento[faturamento["Competencia"] == mes]
        nome_aba = mes.replace("/", "-")
        
        df_mes.to_excel(writer, sheet_name=nome_aba, index=False, startrow=3)
        ws = writer.book[nome_aba]

        # TÍTULO
        ws["A1"] = "Relatório de Faturamento e Impostos"
        ws["A2"] = f"Competência: {mes}"
        
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"].font = Font(size=12)

        # CABEÇALHO
        fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill

        # CONGELAR
        ws.freeze_panes = "A5"

        # FILTRO
        ws.auto_filter.ref = f"A4:J{4 + len(df_mes)}"

        # MOEDA
        colunas_moeda = ["D", "E", "F", "H", "I", "J"]
        
        for col in colunas_moeda:
            for row in range(5, 5 + len(df_mes)):
                ws[f"{col}{row}"].number_format = 'R$ #,##0.00'

        # AJUSTE DE COLUNAS
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            ws.column_dimensions[col_letter].width = max_length + 2

        # TABELA
        ultima_linha = 4 + len(df_mes)
        tabela = Table(
            displayName=f"Tabela_{mes.replace('/', '')}",
            ref=f"A4:J{ultima_linha}"
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

        # DESTAQUE TRIMESTRAL
        mes_num = int(mes.split("/")[0])
        if mes_num in [3, 6, 9, 12]:
            ws["A2"].font = Font(size=12, bold=True)
            ws["A2"].value += " (Fechamento Trimestral)"

print(f"\nArquivo gerado com sucesso: {arquivo_saida}")